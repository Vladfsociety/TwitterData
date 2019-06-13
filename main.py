from openpyxl import load_workbook, Workbook
from openpyxl.writer.excel import save_workbook
import pandas as pd
from datetime import datetime


def GetDataFrame(fileName):
    workBook = load_workbook(fileName)
    sheetData = workBook['Data']
    data = sheetData.values
    cols = next(data)[0:]
    idx = [(r + 2) for r in range(200)]
    df = pd.DataFrame(data, index = idx, columns=cols)
    return df


def UniqueChannels(df):
    uniqueChannelName = {}
    for row in range(2, 200):
        if df.iloc[row][r'twitter Channel'] not in uniqueChannelName:
            uniqueChannelName[df.iloc[row][r'twitter Channel']] = None
    return uniqueChannelName


def UniqueDate(df):
    uniqueDate = {}
    for row in range(2, 200):
        if df.iloc[row]['Date'] not in uniqueDate:
            uniqueDate[df.iloc[row]['Date']] = None
    return uniqueDate


def WriteToExcelFile(resFileName, df):
    try:
        workBook = load_workbook(resFileName)
    except:
        workBook = Workbook()
        for sheet_name in workBook.sheetnames:
            sheet = workBook[sheet_name]
            workBook.remove(sheet)
    uniqueDate = UniqueDate(df)
    uniqueChannelName = UniqueChannels(df)
    WriteCountChannelPerDay(df, workBook, resFileName, uniqueDate, uniqueChannelName)
    WriteRetwitountChannelPerDay(df, workBook, resFileName, uniqueDate, uniqueChannelName)
    WriteFavoritecountChannelPerDay(df, workBook, resFileName, uniqueDate, uniqueChannelName)


def WriteCountChannelPerDay(df, workBook, resFileName, uniqueDate, uniqueChannelName):
    for name in uniqueChannelName:
        newSheet = workBook.create_sheet(r"Postcount " + name)
        dfChannel = df[df[r'twitter Channel'] == name]
        col = ["Date", r"Post count"]
        rowIndex = 1
        newSheet.cell(row=rowIndex, column=1).value = col[0]
        newSheet.cell(row=rowIndex, column=2).value = col[1]
        rowIndex += 1
        for date in uniqueDate:
            dfDateChannel = dfChannel[dfChannel['Date'] == date]
            if dfDateChannel.count()['Date'] == 0:
                newSheet.cell(row=rowIndex, column=1).value = date
                newSheet.cell(row=rowIndex, column=2).value = 0
                rowIndex += 1
                continue
            newSheet.cell(row=rowIndex, column=1).value = dfDateChannel.iloc[0]['Date']
            newSheet.cell(row=rowIndex, column=2).value = dfDateChannel.count()[r'twitter Channel']
            rowIndex += 1
        save_workbook(workBook, resFileName)


def WriteRetwitountChannelPerDay(df, workBook, resFileName, uniqueDate, uniqueChannelName):
    for name in uniqueChannelName:
        newSheet = workBook.create_sheet(r"Retwitcount " + name)
        dfChannel = df[df[r'twitter Channel'] == name]
        col = ["Date", "Retwitcount"]
        rowIndex = 1
        newSheet.cell(row=rowIndex, column=1).value = col[0]
        newSheet.cell(row=rowIndex, column=2).value = col[1]
        rowIndex += 1
        for date in uniqueDate:
            dfDateChannel = dfChannel[dfChannel['Date'] == date]
            if dfDateChannel.count()['Date'] == 0:
                newSheet.cell(row=rowIndex, column=1).value = date
                newSheet.cell(row=rowIndex, column=2).value = 0
                rowIndex += 1
                continue
            sum = 0
            for i in range(dfDateChannel.count()['Retwitcount']):
                sum += dfDateChannel.iloc[i]['Retwitcount']
            newSheet.cell(row=rowIndex, column=1).value = dfDateChannel.iloc[0]['Date']
            newSheet.cell(row=rowIndex, column=2).value = sum
            rowIndex += 1
        save_workbook(workBook, resFileName)


def WriteFavoritecountChannelPerDay(df, workBook, resFileName, uniqueDate, uniqueChannelName):
    for name in uniqueChannelName:
        newSheet = workBook.create_sheet(r"Favoritecount " + name)
        dfChannel = df[df[r'twitter Channel'] == name]
        col = ["Date", "Favoritecount"]
        rowIndex = 1
        newSheet.cell(row=rowIndex, column=1).value = col[0]
        newSheet.cell(row=rowIndex, column=2).value = col[1]
        rowIndex += 1
        for date in uniqueDate:
            dfDateChannel = dfChannel[dfChannel['Date'] == date]
            if dfDateChannel.count()['Date'] == 0:
                newSheet.cell(row=rowIndex, column=1).value = date
                newSheet.cell(row=rowIndex, column=2).value = 0
                rowIndex += 1
                continue
            sum = 0
            for i in range(dfDateChannel.count()['Favoritecount']):
                sum += dfDateChannel.iloc[i]['Favoritecount']
            newSheet.cell(row=rowIndex, column=1).value = dfDateChannel.iloc[0]['Date']
            newSheet.cell(row=rowIndex, column=2).value = sum
            rowIndex += 1
        save_workbook(workBook, resFileName)



fileName = r'DA twitter task (v2).xlsx'
resFileName = r'Кобыльник В.А..xlsx'
df = GetDataFrame(fileName)
WriteToExcelFile(resFileName, df)
